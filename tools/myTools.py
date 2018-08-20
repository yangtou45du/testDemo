#!/usr/bin/python
# -*- coding: cp936 -*-
# -*- coding: encoding -*-
import xlrd,sys
from openpyxl import load_workbook
from openpyxl import Workbook
import cx_Oracle

class Excel():
    def read_it(self, path, index=0):
        data=xlrd.open_workbook(path)
        sheet=data.sheets()[index]
        return sheet
def removeExcel():
    wb=load_workbook("C:\Users\Lenovo\PycharmProjects\\testdemo\\testdata\\test_result.xlsx")
    sheet=wb.active
    wb.remove(sheet)#s删除上一次报告的内容
    wb1=Workbook("C:\Users\Lenovo\PycharmProjects\\testdemo\\testdata\\test_result.xlsx")
    wb1.create_sheet(title="test_result",index=0)#新建一个工作表
    wb1.save("C:\Users\Lenovo\PycharmProjects\\testdemo\\testdata\\test_result.xlsx")
    wb1.close()
def write_log(msg):
    f=open("C:\\Users\\Lenovo\\PycharmProjects\\testdemo\\testdata\\log.txt","a")
    f.write(msg)
    f.close()
def write(list):
    filename="C:\Users\Lenovo\PycharmProjects\\testdemo\\testdata\\test_result.xlsx"
    wb=load_workbook(filename)
    sheet = wb.active # 激活工作页
    sheet.append(list)
    wb.close()
    wb.save("C:\Users\Lenovo\PycharmProjects\\testdemo\\testdata\\test_result.xlsx")

def get_dc(string):
    list=string.split("\n")
    dc={}
    for i in list:
        list1=i.split("=")
        dc[list1[0]]=list1[1]
    return dc

