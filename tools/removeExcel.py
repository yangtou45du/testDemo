#!/usr/bin/python
# -*- coding: cp936 -*-
# -*- coding: encoding -*-
from openpyxl import load_workbook
from openpyxl import Workbook
def removeExcel():
    wb=load_workbook("C:\Users\Lenovo\PycharmProjects\\testdemo\\testdata\\test_result.xlsx")
    sheet=wb.active
    wb.remove(sheet)#s删除上一次报告的内容
    wb1=Workbook("C:\Users\Lenovo\PycharmProjects\\testdemo\\testdata\\test_result.xlsx")
    wb1.create_sheet(title="test_result",index=0)#新建一个工作表
    wb1.save("C:\Users\Lenovo\PycharmProjects\\testdemo\\testdata\\test_result.xlsx")
    wb1.close()



#removeExcel()